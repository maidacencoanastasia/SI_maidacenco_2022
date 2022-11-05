from bs4 import BeautifulSoup
from openpyxl import Workbook
import statistics
from selenium import webdriver

PATH = r'D:/Desktop/chromedriver.exe'
driver = webdriver.Chrome(PATH)

url1 = "https://www.pcgarage.ro/procesoare/filtre/serie-intel-core-i9/"
url2 = "https://www.emag.ro/procesoare/filter/familie-procesor-f2666,intel-core-i9-v29361/c?ref=lst_leftbar_2666_29361"
url3 = "https://altex.ro/procesoare-calculator/cpl/filtru/tip-procesor-7195/intel-core-i9/"


def shop1(url1):
    driver.get(url1)
    # https://www.emag.ro/procesoare/filter/familie-procesor-f2666,intel-core-i9-v29361/c?ref=lst_leftbar_2666_29361
    # https://altex.ro/procesoare-calculator/cpl/filtru/tip-procesor-7195/intel-core-i9/
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    response = soup.find_all("p", {"class": "price"})
    data = []
    for item in response:
        data.append(item.text.strip("RON"))
    names1 = []
    response2 = soup.find_all("h2", {"class": "my-0"})
    for item in response2:
        names1.append(item.text)
    # print(data)
    mod_data = []
    for item in data:
        item = item.replace('.', '')
        item = item.replace(',', '.')
        item = mod_data.append(float(item))
    print(mod_data)
    print(statistics.mean(mod_data))
    return mod_data, names1


def shop2(url2):
    # PATH = r'D:/Desktop/chromedriver.exe'
    # driver = webdriver.Chrome(PATH)
    # driver.get("https://www.emag.ro/procesoare/filter/familie-procesor-f2666,intel-core-i9-v29361/c?ref=lst_leftbar_2666_29361")
    driver.get(url2)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    response = soup.find_all("p", {"class": "product-new-price"})

    data = []

    for item in response:
        data.append((item.text.strip("Lei")))

    response = soup.find_all("a", {"class": "card-v2-title"})

    names2 = []

    for item in response:
        names2.append(item.text)

    # print(data)
    mod_data = []
    for item in data:
        item = item.replace('.', '')
        item = item.replace(',', '.')
        item = mod_data.append(float(item))
    print(mod_data)
    print(statistics.mean(mod_data))
    return mod_data, names2


def shop3(url3):
    # PATH = r'D:/Desktop/chromedriver.exe'
    # driver = webdriver.Chrome(PATH)
    # driver.get("https://altex.ro/procesoare-calculator/cpl/filtru/tip-procesor-7195/intel-core-i9/")
    driver.get(url3)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    response = soup.find_all("span", {"class": "Price-int leading-none"})

    data = []

    for item in response:
        data.append((item.text.strip("lei")))

    response = soup.find_all("h2", {"class": "Product-nameHeading"})

    names3 = []

    for item in response:
        names3.append((item.text.strip("lei")))

    # print(data)
    mod_data = []
    for item in data:
        item = item.replace('.', '')
        item = item.replace(',', '.')
        item = mod_data.append(float(item))
    print(mod_data)
    print(statistics.mean(mod_data))
    return mod_data, names3


print("GARAJ")
extracted_data1, extracted_name1 = shop1(url1)
print("EMAG")
extracted_data2, extracted_name2 = shop2(url2)
print("ALTEX")
extracted_data3, extracted_name3 = shop3(url3)

# mag1 = statistics.mean(extracted_data1)
# mag2 = statistics.mean(extracted_data2)
# mag3 = statistics.mean(extracted_data3)
# print(mag1, mag2, mag3)

wb = Workbook()
#wb.create_sheet('GARAJ')
ws = wb.create_sheet('GARAJ')
ws['A1'] = "GARAJ"
for i, name in enumerate(extracted_data1):
    ws[f"A{i + 2}"] = (name)
for i, name in enumerate(extracted_name1):
    ws[f"B{i + 2}"] = (name)

#wb.create_sheet('EMAG')
#ws = wb.active
ws = wb.create_sheet('EMAG')
ws['A1'] = "EMAG"
for i, name in enumerate(extracted_data2):
    ws[f"A{i + 2}"] = (name)
for i, name in enumerate(extracted_name2):
    ws[f"B{i + 2}"] = (name)

#wb.create_sheet('ALTEX')
# ws = wb.create_sheet('ALTEX')
# ws['F1'] = "ALTEX"
ws = wb.create_sheet('ALTEX')
ws['A1'] = "ALTEX"
for i, name in enumerate(extracted_data3):
    ws[f"A{i + 2}"] = (name)
for i, name in enumerate(extracted_name3):
    ws[f"B{i + 2}"] = (name)

#ws.set_column("B:B",100)
wb.save("data.xlsx")
